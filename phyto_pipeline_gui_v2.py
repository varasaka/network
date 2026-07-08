"""
Phytochemical Network Pharmacology Pipeline (GUI) — v2
=========================================================

Type a plant name -> pick which IMPPAT columns to keep -> the app:

    1. Scrapes IMPPAT for every phytochemical of that plant
       (Compound name, SMILES, Molecular formula, Molecular weight)
    2. Calls SwissADME (via the ToolUniverse library) for each
       compound to get ADME/drug-likeness data
    3. Flags "bioactive" compounds using:
           - Lipinski Rule of Five: <= 1 violation
           - GI absorption: High
           - Bioavailability Score >= 0.55
    4. Calls SwissTargetPrediction (via ToolUniverse) for each
       bioactive compound to get predicted protein targets (human)
    5. Saves everything into one multi-sheet Excel workbook:
           - Compounds            (raw IMPPAT data)
           - ADME_All             (SwissADME results, all compounds)
           - Bioactive_Compounds  (compounds passing the filter)
           - Target_Predictions   (predicted targets for bioactive compounds)

WHY TOOLUNIVERSE INSTEAD OF SCRAPING:
    SwissADME and SwissTargetPrediction's own websites are fragile to
    scrape directly (batch jobs can silently fail, results are written
    asynchronously, HTML structure needs constant babysitting). The
    ToolUniverse project (github.com/mims-harvard/ToolUniverse,
    pip install tooluniverse) already wraps both tools as clean,
    single-molecule API calls with structured JSON output, and handles
    the underlying timing/retry logic internally. We use it here
    instead of maintaining our own scraper for those two steps.

    IMPPAT itself has no such tool available, so Step 1 still uses
    the directly-verified scraper from earlier in this project.

Requirements:
    pip install requests beautifulsoup4 openpyxl tooluniverse --break-system-packages

Run:
    python phyto_pipeline_gui_v2.py

NOTE: The first pipeline run will take an extra 10-15 seconds to load
ToolUniverse's tool registry (600+ tools) — this only happens once.
SwissADME calls take ~5-15s per compound; SwissTargetPrediction calls
take ~30-60s per compound (it runs a real similarity search against
~370,000 known active compounds). For a plant with many bioactive
compounds, Step 4 can take a while — the log window shows live
per-compound progress so it won't look frozen.
"""

import re
import sys
import time
import threading
import urllib.parse
from pathlib import Path
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

# ==================================================================
# Shared HTTP session (used for IMPPAT only)
# ==================================================================

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    )
}
SESSION = requests.Session()
SESSION.headers.update(HEADERS)
REQUEST_TIMEOUT = 60


def get_soup(url: str, **kwargs) -> BeautifulSoup:
    resp = SESSION.get(url, timeout=REQUEST_TIMEOUT, **kwargs)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "html.parser")


# ==================================================================
# STEP 1 — IMPPAT scraping (verified working against real HTML)
# ==================================================================

IMPPAT_BASE = "https://cb.imsc.res.in/imppat"
PLANT_URL_TMPL = f"{IMPPAT_BASE}/phytochemical/{{plant}}"
DETAIL_URL_TMPL = f"{IMPPAT_BASE}/phytochemical-detailedpage/{{cid}}"
PHYSCHEM_URL_TMPL = f"{IMPPAT_BASE}/physicochemicalproperties/{{cid}}"

INCHI_FORMULA_RE = re.compile(r"InChI=1S?/([A-Za-z0-9.]+)/")
IMPPAT_DELAY = 0.4


def fetch_compound_list(plant_name: str, log):
    plant_url = PLANT_URL_TMPL.format(plant=urllib.parse.quote(plant_name))
    soup = get_soup(plant_url)

    table = soup.find("table", id="table_id") or soup.find("table")
    if table is None:
        log(f"No phytochemical table found for '{plant_name}'. "
            f"Check the spelling matches IMPPAT exactly.")
        return []

    compounds = {}
    for row in table.find("tbody").find_all("tr"):
        cells = row.find_all("td")
        if len(cells) < 4:
            continue
        id_cell, name_cell = cells[2], cells[3]
        id_link = id_cell.find("a")
        name_link = name_cell.find("a")
        if not id_link or not name_link:
            continue
        cid = id_link.get_text(strip=True)
        name = name_link.get_text(strip=True)
        if cid and name and cid not in compounds:
            compounds[cid] = name

    info_div = soup.find(id="table_id_info")
    if info_div:
        m = re.search(r"of\s+([\d,]+)\s+entries", info_div.get_text())
        if m:
            total_expected = int(m.group(1).replace(",", ""))
            if len(compounds) < total_expected:
                log(f"Warning: page reports {total_expected} total entries but "
                    f"only {len(compounds)} were found — results may be incomplete.")

    return list(compounds.items())


def extract_smiles_and_inchi(soup: BeautifulSoup):
    smiles, inchi = "", ""
    for strong in soup.find_all("strong"):
        label = strong.get_text(strip=True)
        if label == "SMILES:":
            text_tag = strong.find_next("text")
            if text_tag:
                smiles = text_tag.get_text(strip=True)
        elif label == "InChI:":
            text_tag = strong.find_next("text")
            if text_tag:
                inchi = text_tag.get_text(strip=True)
    return smiles, inchi


def formula_from_inchi(inchi: str) -> str:
    if not inchi:
        return ""
    m = INCHI_FORMULA_RE.search(inchi)
    return m.group(1) if m else ""


def extract_mw(soup: BeautifulSoup) -> str:
    for row in soup.find_all("tr"):
        cells = row.find_all("td")
        if not cells:
            continue
        label = cells[0].get_text(strip=True).lower()
        if "molecular weight" in label:
            value_cell = cells[-1].get_text(strip=True)
            m = re.search(r"[\d]+\.?\d*", value_cell)
            if m:
                return m.group(0)
    return ""


def scrape_imppat_compound(cid: str, name: str) -> dict:
    row = {"Compound_ID": cid, "Name": name, "SMILES": "", "Formula": "", "MW_g_per_mol": ""}
    try:
        detail_soup = get_soup(DETAIL_URL_TMPL.format(cid=cid))
        smiles, inchi = extract_smiles_and_inchi(detail_soup)
        row["SMILES"] = smiles
        row["Formula"] = formula_from_inchi(inchi)
    except Exception:
        pass
    time.sleep(IMPPAT_DELAY)
    try:
        phys_soup = get_soup(PHYSCHEM_URL_TMPL.format(cid=cid))
        row["MW_g_per_mol"] = extract_mw(phys_soup)
    except Exception:
        pass
    time.sleep(IMPPAT_DELAY)
    return row


# ==================================================================
# ToolUniverse setup (used for Steps 2 and 4)
# ==================================================================

_tu_instance = None


def get_tooluniverse(log):
    """Lazily create and load the ToolUniverse registry (once)."""
    global _tu_instance
    if _tu_instance is None:
        log("Loading ToolUniverse tool registry (one-time, ~10-15s)...")
        from tooluniverse import ToolUniverse
        _tu_instance = ToolUniverse()
        _tu_instance.load_tools()
        log("ToolUniverse ready.")
    return _tu_instance


# ==================================================================
# STEP 2 — SwissADME via ToolUniverse
# ==================================================================

def run_swissadme_one(tu, smiles: str, name: str, log, max_retries: int = 3) -> dict:
    """
    Calls SwissADME_calculate_adme for a single compound and flattens
    the nested JSON response into a single dict for easy Excel export.
    Returns {} on failure (logged).

    Retries with backoff on failure: SwissADME's server appears to
    rate-limit rapid-fire requests (observed: fast successive calls
    start failing near-instantly after ~15-20 compounds, and can stay
    blocked across app restarts for a short window). A real invalid
    SMILES will keep failing on retry too, so retrying is safe either
    way — it just costs a little extra time for genuinely bad SMILES.
    """
    last_error = None

    for attempt in range(1, max_retries + 1):
        try:
            result = tu.run({
                "name": "SwissADME_calculate_adme",
                "arguments": {
                    "operation": "calculate_adme",
                    "smiles": smiles,
                    "molecule_name": name,
                },
            })
        except Exception as e:
            last_error = str(e)
            result = None

        if isinstance(result, dict) and "error" not in result:
            flat = {
                "Compound_Name": result.get("molecule_name") or name,
                "SMILES": result.get("canonical_smiles") or smiles,
            }
            for section in ("physicochemical", "lipophilicity", "water_solubility",
                             "pharmacokinetics", "druglikeness", "medicinal_chemistry"):
                sub = result.get(section) or {}
                if isinstance(sub, dict):
                    for k, v in sub.items():
                        flat[f"{section}.{k}"] = v
            return flat

        last_error = (result.get("error") if isinstance(result, dict) else result) or last_error

        if attempt < max_retries:
            backoff = 8.0 * attempt  # 8s, then 16s
            log(f"    [warn] SwissADME failed for '{name}' (attempt "
                f"{attempt}/{max_retries}): {last_error}. "
                f"Waiting {backoff:.0f}s before retry (likely rate limiting)...")
            time.sleep(backoff)

    log(f"    [error] SwissADME gave up on '{name}' after {max_retries} "
        f"attempts: {last_error}")
    return {}


# ==================================================================
# STEP 3 — Bioactivity filter
# ==================================================================

def is_bioactive(adme_row: dict) -> bool:
    """
    Bioactivity rule:
        - Lipinski Rule of Five: <= 1 violation
        - GI absorption: High
        - Bioavailability Score >= 0.55
    Reads the flattened field names produced by run_swissadme_one().
    """
    lipinski_raw = adme_row.get("druglikeness.lipinski_violations")
    gi_raw = adme_row.get("pharmacokinetics.gi_absorption") or ""
    bioav_raw = adme_row.get("druglikeness.bioavailability_score")

    try:
        lipinski_violations = int(str(lipinski_raw).strip())
    except (TypeError, ValueError):
        lipinski_violations = 99  # unknown -> fail safe

    gi_high = str(gi_raw).strip().lower() == "high"

    try:
        bioav_score = float(bioav_raw)
    except (TypeError, ValueError):
        bioav_score = 0.0

    return lipinski_violations <= 1 and gi_high and bioav_score >= 0.55


# ==================================================================
# STEP 4 — SwissTargetPrediction via ToolUniverse
# ==================================================================

def run_swisstargetprediction_one(tu, smiles: str, name: str, log) -> list:
    """
    Calls SwissTargetPrediction_predict for a single compound.
    Returns a list of dicts, one per predicted target.
    """
    try:
        result = tu.run({
            "name": "SwissTargetPrediction_predict",
            "arguments": {
                "operation": "predict",
                "smiles": smiles,
                "organism": "Homo_sapiens",
            },
        })
    except Exception as e:
        log(f"    [error] SwissTargetPrediction call failed for '{name}': {e}")
        return []

    if not isinstance(result, dict) or "error" in result:
        log(f"    [error] SwissTargetPrediction returned an error for "
            f"'{name}': {result.get('error') if isinstance(result, dict) else result}")
        return []

    targets = result.get("targets") or []
    rows = []
    for t in targets:
        rows.append({
            "Compound": name,
            "SMILES": smiles,
            "Target": t.get("target_name", ""),
            "Gene Symbol": t.get("gene_symbol", ""),
            "Uniprot ID": t.get("uniprot_id", ""),
            "ChEMBL ID": t.get("chembl_id", ""),
            "Target Class": t.get("target_class", ""),
            "Probability": t.get("probability", ""),
            "Known actives (3D)": t.get("known_actives_3d", ""),
            "Known actives (2D)": t.get("known_actives_2d", ""),
        })
    return rows


# ==================================================================
# Excel writer
# ==================================================================

HEADER_FILL = PatternFill(start_color="1F4E3D", end_color="1F4E3D", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws, ncols):
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, ncols, max_width=60):
    for col_idx in range(1, ncols + 1):
        letter = get_column_letter(col_idx)
        best = 10
        for cell in ws[letter]:
            if cell.value:
                best = max(best, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = best


def write_pipeline_excel(out_path: Path, compounds: list, columns_selected: dict,
                          adme_rows: list, bioactive_rows: list, target_rows: list):
    wb = Workbook()

    # --- Sheet 1: Compounds (respecting selected columns) ---
    ws1 = wb.active
    ws1.title = "Compounds"
    headers1 = ["Compound ID"]
    if columns_selected.get("name", True):
        headers1.append("Name")
    if columns_selected.get("smiles", True):
        headers1.append("SMILES")
    if columns_selected.get("formula", True):
        headers1.append("Formula")
    if columns_selected.get("mw", True):
        headers1.append("MW (g/mol)")
    ws1.append(headers1)

    for row in compounds:
        line = [row.get("Compound_ID", "")]
        if columns_selected.get("name", True):
            line.append(row.get("Name", ""))
        if columns_selected.get("smiles", True):
            line.append(row.get("SMILES", ""))
        if columns_selected.get("formula", True):
            line.append(row.get("Formula", ""))
        if columns_selected.get("mw", True):
            line.append(row.get("MW_g_per_mol", ""))
        ws1.append(line)
    _style_header(ws1, len(headers1))
    _autosize(ws1, len(headers1))

    # --- Sheet 2: ADME_All ---
    ws2 = wb.create_sheet("ADME_All")
    if adme_rows:
        headers2 = list(adme_rows[0].keys())
        ws2.append(headers2)
        for row in adme_rows:
            ws2.append([row.get(h, "") for h in headers2])
        _style_header(ws2, len(headers2))
        _autosize(ws2, len(headers2))
    else:
        ws2.append(["No ADME data retrieved."])

    # --- Sheet 3: Bioactive_Compounds ---
    ws3 = wb.create_sheet("Bioactive_Compounds")
    if bioactive_rows:
        headers3 = list(bioactive_rows[0].keys())
        ws3.append(headers3)
        for row in bioactive_rows:
            ws3.append([row.get(h, "") for h in headers3])
        _style_header(ws3, len(headers3))
        _autosize(ws3, len(headers3))
    else:
        ws3.append(["No compounds passed the bioactivity filter."])

    # --- Sheet 4: Target_Predictions ---
    ws4 = wb.create_sheet("Target_Predictions")
    if target_rows:
        headers4 = list(target_rows[0].keys())
        ws4.append(headers4)
        for row in target_rows:
            ws4.append([row.get(h, "") for h in headers4])
        _style_header(ws4, len(headers4))
        _autosize(ws4, len(headers4))
    else:
        ws4.append(["No target predictions retrieved."])

    wb.save(out_path)


# ==================================================================
# GUI
# ==================================================================

class PipelineApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Phytochemical Network Pharmacology Pipeline")
        self.geometry("760x600")
        self.minsize(680, 520)

        self.outdir = Path.cwd()

        # --- input row ---
        top = ttk.Frame(self, padding=12)
        top.pack(fill="x")

        ttk.Label(top, text="Plant name:").pack(side="left")
        self.entry = ttk.Entry(top)
        self.entry.pack(side="left", fill="x", expand=True, padx=8)
        self.entry.bind("<Return>", lambda e: self.on_run())
        self.entry.focus_set()

        self.run_btn = ttk.Button(top, text="Run Pipeline", command=self.on_run)
        self.run_btn.pack(side="left")

        # --- column selection ---
        col_frame = ttk.LabelFrame(self, text="Columns to include for Compounds sheet", padding=10)
        col_frame.pack(fill="x", padx=12, pady=(0, 8))

        self.var_name = tk.BooleanVar(value=True)
        self.var_smiles = tk.BooleanVar(value=True)
        self.var_formula = tk.BooleanVar(value=True)
        self.var_mw = tk.BooleanVar(value=True)

        ttk.Checkbutton(col_frame, text="Compound name", variable=self.var_name).pack(side="left", padx=6)
        ttk.Checkbutton(col_frame, text="SMILES", variable=self.var_smiles, state="disabled").pack(side="left", padx=6)
        ttk.Checkbutton(col_frame, text="Molecular formula", variable=self.var_formula).pack(side="left", padx=6)
        ttk.Checkbutton(col_frame, text="Molecular weight", variable=self.var_mw).pack(side="left", padx=6)
        ttk.Label(col_frame, text="(SMILES always included — required for the pipeline)",
                  foreground="#666").pack(side="left", padx=10)

        # --- pipeline toggle ---
        toggle_frame = ttk.Frame(self, padding=(12, 0))
        toggle_frame.pack(fill="x")
        self.var_run_downstream = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            toggle_frame,
            text="Also run SwissADME + bioactivity filter + SwissTargetPrediction (uncheck for IMPPAT data only)",
            variable=self.var_run_downstream,
        ).pack(side="left")

        # --- output folder row ---
        folder_row = ttk.Frame(self, padding=(12, 6))
        folder_row.pack(fill="x")
        self.folder_label = ttk.Label(folder_row, text=f"Save to: {self.outdir}")
        self.folder_label.pack(side="left")
        ttk.Button(folder_row, text="Change...", command=self.choose_folder).pack(side="right")

        # --- log box ---
        log_frame = ttk.Frame(self, padding=12)
        log_frame.pack(fill="both", expand=True)

        self.log_box = tk.Text(log_frame, wrap="word", state="disabled", height=20)
        self.log_box.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(log_frame, command=self.log_box.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_box["yscrollcommand"] = scrollbar.set

        # --- status bar ---
        self.status = ttk.Label(self, text="Ready.", padding=(12, 4), anchor="w")
        self.status.pack(fill="x")

    def log(self, msg: str):
        self.log_box.configure(state="normal")
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_box.insert("end", f"[{timestamp}] {msg}\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def choose_folder(self):
        chosen = filedialog.askdirectory(initialdir=str(self.outdir))
        if chosen:
            self.outdir = Path(chosen)
            self.folder_label.config(text=f"Save to: {self.outdir}")

    def on_run(self):
        plant_name = self.entry.get().strip()
        if not plant_name:
            messagebox.showwarning("Missing input", "Please type a plant name first.")
            return

        self.run_btn.config(state="disabled")
        self.status.config(text=f"Running pipeline for '{plant_name}'...")
        self.log(f"=== Starting pipeline for '{plant_name}' ===")

        thread = threading.Thread(target=self.run_pipeline, args=(plant_name,), daemon=True)
        thread.start()

    def run_pipeline(self, plant_name: str):
        def log(msg):
            self.after(0, self.log, msg)

        def status(msg):
            self.after(0, self.status.config, {"text": msg})

        try:
            # --- Step 1: IMPPAT ---
            status("Step 1/4: Fetching compound list from IMPPAT...")
            compound_ids = fetch_compound_list(plant_name, log)
            log(f"Found {len(compound_ids)} unique compounds.")

            compounds = []
            for i, (cid, name) in enumerate(compound_ids, 1):
                compounds.append(scrape_imppat_compound(cid, name))
                log(f"  [{i}/{len(compound_ids)}] {cid}  {name}")
                status(f"Step 1/4: IMPPAT {i}/{len(compound_ids)}: {name}")

            columns_selected = {
                "name": self.var_name.get(),
                "smiles": True,
                "formula": self.var_formula.get(),
                "mw": self.var_mw.get(),
            }

            adme_rows, bioactive_rows, target_rows = [], [], []

            if self.var_run_downstream.get() and compounds:
                tu = get_tooluniverse(log)

                # --- Step 2: SwissADME (one compound at a time via ToolUniverse) ---
                status("Step 2/4: Running SwissADME...")
                log("--- Step 2: SwissADME (via ToolUniverse) ---")
                usable = [c for c in compounds if c["SMILES"]]
                for i, c in enumerate(usable, 1):
                    log(f"  [{i}/{len(usable)}] SwissADME: {c['Name']}...")
                    status(f"Step 2/4: SwissADME {i}/{len(usable)}: {c['Name']}")
                    flat = run_swissadme_one(tu, c["SMILES"], c["Name"], log)
                    if flat:
                        adme_rows.append(flat)
                    # Pace requests to avoid tripping SwissADME's rate
                    # limiting (observed: rapid-fire calls start failing
                    # after ~15-20 compounds without this).
                    time.sleep(3.0)
                log(f"SwissADME returned data for {len(adme_rows)} of {len(usable)} compounds.")

                # --- Step 3: Bioactivity filter ---
                status("Step 3/4: Applying bioactivity filter...")
                log("--- Step 3: Bioactivity filter "
                    "(Lipinski <=1 violation, GI absorption High, Bioavailability >= 0.55) ---")
                for row in adme_rows:
                    if is_bioactive(row):
                        bioactive_rows.append(row)
                log(f"{len(bioactive_rows)} of {len(adme_rows)} compounds passed the filter.")

                # --- Step 4: SwissTargetPrediction ---
                status("Step 4/4: Predicting targets for bioactive compounds...")
                log("--- Step 4: SwissTargetPrediction (via ToolUniverse) ---")
                for i, row in enumerate(bioactive_rows, 1):
                    smiles = row.get("SMILES", "")
                    compound_name = row.get("Compound_Name", "Unknown")
                    log(f"  [{i}/{len(bioactive_rows)}] Predicting targets for "
                        f"{compound_name} (can take 30-60s)...")
                    status(f"Step 4/4: Target prediction {i}/{len(bioactive_rows)}: {compound_name}")
                    targets = run_swisstargetprediction_one(tu, smiles, compound_name, log)
                    log(f"    -> {len(targets)} predicted targets.")
                    target_rows.extend(targets)
            else:
                log("Downstream pipeline (ADME/target prediction) skipped.")

            # --- Save Excel ---
            safe_name = re.sub(r"\s+", "_", plant_name.strip())
            out_path = self.outdir / f"{safe_name}_pipeline_results.xlsx"
            write_pipeline_excel(out_path, compounds, columns_selected,
                                  adme_rows, bioactive_rows, target_rows)

            log(f"Saved: {out_path}")
            status(f"Done — results saved to {out_path.name}")
            self.after(0, lambda: messagebox.showinfo(
                "Done",
                f"Pipeline complete for '{plant_name}'.\n\n"
                f"Compounds: {len(compounds)}\n"
                f"ADME results: {len(adme_rows)}\n"
                f"Bioactive compounds: {len(bioactive_rows)}\n"
                f"Target predictions: {len(target_rows)}\n\n"
                f"Saved to:\n{out_path}"
            ))
        except requests.exceptions.RequestException as e:
            log(f"Network error: {e}")
            self.after(0, lambda: messagebox.showerror("Network error", str(e)))
            status("Failed — see log.")
        except Exception as e:
            log(f"Error: {e}")
            self.after(0, lambda: messagebox.showerror("Error", str(e)))
            status("Failed — see log.")
        finally:
            self.after(0, self.run_btn.config, {"state": "normal"})


if __name__ == "__main__":
    app = PipelineApp()
    app.mainloop()
